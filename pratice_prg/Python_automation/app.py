import openpyxl as xl
from openpyxl.chart import BarChart, Reference 
def PriceCorrection(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # Remove the dollar sign and convert to float
        price_value = str(cell.value).replace('$', '')
        corrected_price = float(price_value) * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        values = Reference(sheet, min_col=4, min_row=1, max_col=4, max_row=sheet.max_row)
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, "E5")
    wb.save(filename)
    print(f"Price correction completed and saved to {filename}")